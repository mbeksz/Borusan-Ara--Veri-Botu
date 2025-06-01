import tkinter as tk
from tkinter import messagebox
import threading
import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager

MARKALAR_ORJINAL = [
    "Alfa Romeo", "Audi", "BMW", "Chevrolet", "Citroen", "Cupra", "Defender", "Discovery",
    "DS Automobiles", "Fiat", "Ford", "Honda", "Hyundai", "Jaguar", "Jeep", "Kia",
    "Land Rover", "Mercedes - Benz", "MG", "MINI", "Nissan", "Opel", "Peugeot", "Porsche",
    "Range Rover", "Renault", "Seat", "Skoda", "Skywell", "Subaru", "Suzuki", "Tesla",
    "Togg", "Toyota", "Volkswagen", "Volvo"
]

current_markalar = MARKALAR_ORJINAL.copy()
show_browser = False  # Kullanıcı tercihini burada saklayacağız

def to_url_brand(name):
    return name.lower().replace(" ", "-")

def scrape_data(selected_brands):
    all_data = []
    brands = [to_url_brand(b) for b in selected_brands]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(os.getcwd(), f"arabalar_borusan_{timestamp}.xlsx")

    chrome_options = Options()
    if not show_browser:
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)

    for brand in brands:
        base_url = f"https://borusannext.com/araba-al/{brand}"
        previous_car_ids = set()
        page = 1
        status_text.set(f"🔎 '{brand.upper()}' aranıyor...")

        while True:
            url = f"{base_url}?pageNumber={page}" if page > 1 else base_url
            driver.get(url)
            time.sleep(2)

            car_elements = driver.find_elements(By.CSS_SELECTOR, "div.vehicle-card")
            if not car_elements:
                break

            # Her sayfa başında tablo başlığı ve ayraç yazdır
            print(f"\n--- {brand.upper()} | Sayfa {page} ---")
            print(f"| {'Marka':<10} | {'Model':<10} | {'Alt Model':<30} | {'Yıl':<4} | {'KM':<10} | {'Yakıt':<10} | {'Vites':<10} | {'Fiyat':<15} | {'Link':<15} | ")
            print("-" * 120)

            for car in car_elements:
                try:
                    href = car.find_element(By.CSS_SELECTOR, "a").get_attribute("href")
 
                    car_id = href.split("/")[-1]
                    if car_id in previous_car_ids:
                        continue
                    previous_car_ids.add(car_id)

                    brand_model_element = car.find_element(By.CSS_SELECTOR, ".vehicle-card-title a")
                    brand_model = brand_model_element.text.strip().split(" ", 1)
                    brand_name = brand_model[0]
                    model = brand_model[1] if len(brand_model) > 1 else ""

                    sub_model = car.find_element(By.CSS_SELECTOR, ".vehicle-card-description").text.strip()

                    specs = car.find_elements(By.CSS_SELECTOR, ".vehicle-card-first-content .grid > div")
                    year = specs[0].text.strip() if len(specs) > 0 else ""
                    mileage = specs[1].text.strip().replace(" Km", "").replace(".", "").strip() if len(specs) > 1 else ""
                    fuel = specs[2].text.strip() if len(specs) > 2 else ""
                    transmission = specs[3].text.strip() if len(specs) > 3 else ""

                    price = car.find_element(By.CSS_SELECTOR, ".vehicle-card-price-text").text.strip()

                    # Tablo satırı olarak yazdır
                    print(f"| {brand_name:<10} | {model:<10} | {sub_model:<30} | {year:<4} | {mileage:<10} | {fuel:<10} | {transmission:<10} | {price:<15} | {href:<15} |")

                    all_data.append({
                        "Marka": brand_name,
                        "Model": model,
                        "Alt Model": sub_model,
                        "Yıl": year,
                        "KM": mileage,
                        "Yakıt": fuel,
                        "Vites": transmission,
                        "Fiyat": price,
                        "Link": href
                    })

                except Exception as e:
                    print("❌ Veri çekilemedi:", e)

            page += 1

    driver.quit()

    df = pd.DataFrame(all_data)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    col_count = ws.max_column
    row_count = ws.max_row
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    cell_range = f"A1:{alphabet[col_count - 1]}{row_count}"

    table = Table(displayName="Veriler", ref=cell_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(output_path)

    status_text.set(f"✅ Dosya kaydedildi: {output_path}")
    open_button.config(command=lambda: os.startfile(output_path))
    open_button.config(state="normal")

def baslat():
    global show_browser
    selections = listbox.curselection()
    if not selections:
        messagebox.showwarning("Uyarı", "Lütfen en az bir marka seçiniz.")
        return

    response = messagebox.askyesno("Tarayıcı Görünsün mü?", "Botun Google üzerinden gezmesini görmek istiyor musunuz?")
    show_browser = response  # Evetse True, Hayırsa False

    selected_brands = [listbox.get(i) for i in selections]
    start_button.config(state="disabled")
    clear_button.config(state="disabled")
    select_all_button.config(state="disabled")
    listbox.config(state="disabled")
    open_button.config(state="disabled")
    status_text.set("🔄 İşlem başladı...")
    threading.Thread(target=worker_thread, args=(selected_brands,), daemon=True).start()

def worker_thread(selected_brands):
    try:
        scrape_data(selected_brands)
    except Exception as e:
        messagebox.showerror("Hata", f"Hata oluştu: {e}")
    finally:
        start_button.config(state="normal")
        clear_button.config(state="normal")
        select_all_button.config(state="normal")
        listbox.config(state="normal")

def sifirla():
    listbox.selection_clear(0, tk.END)
    status_text.set("✅ Seçim sıfırlandı.")
    open_button.config(state="disabled")

def tumunu_sec():
    listbox.select_set(0, tk.END)
    status_text.set("✅ Tüm markalar seçildi.")

def filtrele(*args):
    arama = search_var.get().lower()

    selected_items = [listbox.get(i) for i in listbox.curselection()]

    filtered = [m for m in MARKALAR_ORJINAL if arama in m.lower()]

    for selected in selected_items:
        if selected not in filtered:
            filtered.append(selected)

    listbox.delete(0, tk.END)

    for marka in filtered:
        listbox.insert(tk.END, marka)

    for idx, marka in enumerate(filtered):
        if marka in selected_items:
            listbox.selection_set(idx)



# --- ARAYÜZ ---

root = tk.Tk()
root.title("Borusan Güncel Araçlar Botu")
root.geometry("430x620")
root.resizable(False, False)

tk.Label(root, text="🔎 Marka Arama:", font=("Arial", 11)).pack(pady=(10, 0))
search_var = tk.StringVar()
search_var.trace_add("write", filtrele)
search_entry = tk.Entry(root, textvariable=search_var, width=40)
search_entry.pack(pady=(0, 5))

tk.Label(root, text="📋 Markaları seç", font=("Arial", 10, "bold")).pack()

listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, width=40, height=20, font=("Arial", 10))
for marka in MARKALAR_ORJINAL:
    listbox.insert(tk.END, marka)
listbox.pack(pady=(0, 10))


button_frame = tk.Frame(root)
button_frame.pack(pady=5)

start_button = tk.Button(button_frame, text="▶ Verileri Çek", command=baslat, width=15, bg="green", fg="white")
start_button.grid(row=0, column=0, padx=5)

clear_button = tk.Button(button_frame, text="⛔ Seçimi Sıfırla", command=sifirla, width=15)
clear_button.grid(row=0, column=1, padx=5)

select_all_button = tk.Button(root, text="✅ Tümünü Seç", command=tumunu_sec, width=32)
select_all_button.pack(pady=(5, 5))

open_button = tk.Button(root, text="📂 Dosyayı Aç", state="disabled", width=32)
open_button.pack(pady=(5, 5))

status_text = tk.StringVar()
status_label = tk.Label(root, textvariable=status_text, fg="blue", wraplength=400, justify="center")
status_label.pack(pady=10)

root.mainloop()
