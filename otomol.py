from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Excel'e yazılacak verileri tutan liste
tum_veriler = []

# Terminalden marka isimlerini al (virgülle ayrılmış)
markalar_input = input("Markaları giriniz (örnek: audi, volvo, bmw): ").strip().lower()
markalar = [marka.strip() for marka in markalar_input.split(",")]

# ChromeDriver yolu
service = Service("c:/Users/LENOVO/Documents/chromedriver-win64/chromedriver-win64/chromedriver.exe")
driver = webdriver.Chrome(service=service)

for brand in markalar:
    base_url = f"https://www.otomol.com/{brand}"
    previous_car_ids = set()
    page = 1

    print(f"\n📦 '{brand.upper()}' markası işleniyor...")
    print(f"\n| {'Marka':<10} | {'Model':<10} | {'Alt Model':<30} | {'Yıl':<4} | {'KM':<10} | {'Yakıt':<10} | {'İl':<10} | {'Fiyat':<15} |")
    print("-" * 120)

    while True:
        url = f"{base_url}?p={page}" if page > 1 else base_url
        driver.get(url)
        time.sleep(2)

        arabalar = driver.find_elements(By.CSS_SELECTOR, "tr.product")

        if not arabalar:
            print(f"Sayfa {page} boş, döngü bitiriliyor.")
            break

        current_car_ids = set()
        for araba in arabalar:
            try:
                car_id = araba.get_attribute("item-stokid")
                if car_id in previous_car_ids:
                    continue  # Aynı arabaysa atla
                current_car_ids.add(car_id)

                marka = araba.get_attribute("item-category")
                model = araba.find_element(By.CSS_SELECTOR, ".modelAd").text.strip()
                alt_model = araba.find_element(By.CSS_SELECTOR, ".altModelAd").text.strip()
                yil = araba.find_element(By.CSS_SELECTOR, ".cModelYil").text.strip()
                km = araba.find_element(By.CSS_SELECTOR, ".cAracKm").text.strip()
                yakit = araba.find_element(By.CSS_SELECTOR, ".cYakitTur").text.strip()
                sehir = araba.find_element(By.CSS_SELECTOR, ".cBayi").text.strip()
                fiyat = araba.find_element(By.CSS_SELECTOR, ".cAracFiyat").text.strip()

                print(f"{marka:<8} | {model:<8} | {alt_model:<30} | {yil:<4} | {km:<10} | {yakit:<8} | {sehir:<10} | {fiyat}")

                tum_veriler.append({
                    "Marka": marka,
                    "Model": model,
                    "Alt Model": alt_model,
                    "Yıl": yil,
                    "KM": km,
                    "Yakıt": yakit,
                    "İl": sehir,
                    "Fiyat": fiyat
                })
            except Exception as e:
                print("Hata:", e)

        if not current_car_ids or current_car_ids.issubset(previous_car_ids):
            print("Yeni araba bulunamadı, döngü sonlandırılıyor.")
            break

        previous_car_ids.update(current_car_ids)
        page += 1

driver.quit()

# Excel'e yaz
excel_dosya = "arabalar.xlsx"
df = pd.DataFrame(tum_veriler)
df.to_excel(excel_dosya, index=False)

# Format as Table işlemi (openpyxl ile)
wb = load_workbook(excel_dosya)
ws = wb.active

tablo_adi = "ArabalarTablosu"
satir_sayisi = ws.max_row
sutun_sayisi = ws.max_column
alfabe = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
alan = f"A1:{alfabe[sutun_sayisi - 1]}{satir_sayisi}"

table = Table(displayName=tablo_adi, ref=alan)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

wb.save(excel_dosya)

print(f"\n✅ Veriler '{excel_dosya}' dosyasına tablo formatında kaydedildi.")
